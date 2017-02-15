#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import sys
from openpyxl import load_workbook

CELL_DICT = {
    'CVE_2013_7446': 23,
    'CVE_2014_3153': 24,
    'CVE_2015_1528': 25,
    'CVE_2015_1538': 26,
    'CVE_2015_1805': 27,
    'CVE_2015_2686': 28,
    'CVE_2015_3288': 29,
    'CVE_2015_3636': 30,
    'CVE_2015_3824': 31,
    'CVE_2015_3875': 32,
    'CVE_2015_6602': 33,
    'CVE_2015_6612': 34,
    'CVE_2015_6616': 35,
    'CVE_2015_6636': 36,
    'CVE_2015_6637': 37,
    'CVE_2015_6640': 38,
    'CVE_2015_8966': 39,
    'CVE_2016_0728': 40,
    'CVE_2016_0758': 41,
    'CVE_2016_0803': 42,
    'CVE_2016_0804': 43,
    'CVE_2016_0805': 44,
    'CVE_2016_0807': 45,
    'CVE_2016_0810': 46,
    'CVE_2016_0811': 47,
    'CVE_2016_0815': 48,
    'CVE_2016_0816': 49,
    'CVE_2016_0818': 50,
    'CVE_2016_0819': 51,
    'CVE_2016_0826': 52,
    'CVE_2016_0827': 53,
    'CVE_2016_0828': 54,
    'CVE_2016_0829': 55,
    'CVE_2016_0835': 56,
    'CVE_2016_0836': 57,
    'CVE_2016_0837': 58,
    'CVE_2016_0838': 59,
    'CVE_2016_0839': 60,
    'CVE_2016_0840': 61,
    'CVE_2016_0841': 62,
    'CVE_2016_0843': 63,
    'CVE_2016_0844': 64,
    'CVE_2016_0846': 65,
    'CVE_2016_1621': 66,
    'CVE_2016_2108': 67,
    'CVE_2016_2428': 68,
    'CVE_2016_2430': 69,
    'CVE_2016_2503': 70,
    'CVE_2016_2505': 71,
    'CVE_2016_2506': 72,
    'CVE_2016_2507': 73,
    'CVE_2016_2508': 74,
    'CVE_2016_3741': 75,
    'CVE_2016_3742': 76,
    'CVE_2016_3743': 77,
    'CVE_2016_3768': 78,
    'CVE_2016_3819': 79,
    'CVE_2016_3820': 80,
    'CVE_2016_3821': 81,
    'CVE_2016_3840': 82,
    'CVE_2016_3841': 83,
    'CVE_2016_3843': 84,
    'CVE_2016_3857': 85,
    'CVE_2016_3861': 86,
    'CVE_2016_3862': 87,
    'CVE_2016_5195': 88,
    'CVE_2016_5340': 89,
    'CVE_2016_6699': 90,
    'CVE_2016_6728': 91,
    'CVE_2016_6737': 92,
    'CVE_2016_7117': 93,
    'CVE_2016_7911': 94,
    'CVE_2016_8434': 95,
    'CVE_2016_9120': 96,
}

def usage():
    print("[Usage]: python grep_cve.py COLUMN")

def isExist(file_name):
    if os.path.isfile(file_name):
        pass
    else:
        print("Error: The file %s is not exist!" % file_name)
        sys.exit(1)

if len(sys.argv) != 2:
    usage()
    sys.exit(1)

isExist('mario_log.txt')
isExist('device_prop.txt')
isExist('kernel_info.txt')

wb = load_workbook('/home/acusp/Desktop/test.xlsx')
ws = wb['System Vulnerability Raw Data']
COL = ws[sys.argv[1]]

# ======================
# deal device_prop.txt
# ======================
f = open('device_prop.txt', 'r', encoding='utf-8')
devices_info = f.read()
f.close()

info_list = re.findall('ro.product.manufacturer]: \[(.*)\]', devices_info)
info_list.append(''.join(re.findall('ro.product.model]: \[(.*)\]', devices_info)))
info_list.append(''.join(re.findall('ro.product.name]: \[(.*)\]', devices_info)))
info_list.append(''.join(re.findall('ro.board.platform]: \[(.*)\]', devices_info)))
info_list.append(''.join(re.findall('ro.build.version.release]: \[(.*)\]', devices_info)))
info_list.append(''.join(re.findall('ro.build.version.base_os]: \[(.*)\]', devices_info)))
info_list.append(''.join(re.findall('ro.build.version.sdk]: \[(.*)\]', devices_info)))
info_list.append(''.join(re.findall('ro.build.version.incremental]: \[(.*)\]', devices_info)))
info_list.append(''.join(re.findall('ro.build.version.security_patch]: \[(.*)\]', devices_info)))
info_list.append(''.join(re.findall('ro.build.fingerprint]: \[(.*)\]', devices_info)))
info_list.append(''.join(re.findall('ro.build.date.utc]: \[(.*)\]', devices_info)))

for i in range(len(info_list)):
    row = i + 3
    rst = info_list[i]
    COL[row].value = rst

COL[1].value = COL[3].value
COL[2].value = COL[4].value

# ======================
# create info.txt
# ======================
info = []
info.append(''.join(re.findall('.*ro.product.brand\].*', devices_info)) + '\n')
info.append(''.join(re.findall('.*ro.product.model\].*', devices_info)))
f = open('../info.txt', 'w')
f.writelines(info)
f.close()

# ======================
# deal kernel_info.txt
# ======================
f = open('kernel_info.txt', 'r', encoding='utf-8')
kernel_info = f.read()
f.close()

kernel_str = ''.join(re.findall('kernel_info: (.*)', kernel_info))
COL[14].value = kernel_str

# ======================
# deal mario_log.txt
# ======================
f = open('mario_log.txt', 'r', encoding='utf-8')
raw_log = f.read()
f.close()

pattern_cve = r"case: (.*) result:"
pattern_rst = r"case:.*result: (.*)"
cve_list = re.findall(pattern_cve, raw_log)
rst_list = re.findall(pattern_rst, raw_log)

for i in range(len(cve_list)):
    row = CELL_DICT[cve_list[i]]-1
    rst = rst_list[i]
    COL[row].value = rst

wb.save('/home/acusp/Desktop/test.xlsx')
